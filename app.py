# -*- coding: utf-8 -*-
# ============================================================
# Painel de Qualidade — VELOX (multi-meses)
# ============================================================

import os, io, json, re, unicodedata, calendar
from datetime import datetime, date
from typing import Tuple, Optional

import streamlit as st
import pandas as pd
import numpy as np
import altair as alt

import gspread
from oauth2client.service_account import ServiceAccountCredentials
from dateutil.relativedelta import relativedelta

# Drive API (fallback XLSX)
from google.oauth2 import service_account as gcreds
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


# ------------------ CONFIG BÁSICA ------------------
st.set_page_config(page_title="Painel de Qualidade — VELOX", layout="wide")
st.title("🎯 Painel de Qualidade — VELOX")

st.markdown(
    """
<style>
.card-wrap{display:flex;gap:16px;flex-wrap:wrap;margin:12px 0 6px;}
.card{background:#f7f7f9;border-radius:12px;box-shadow:0 1px 4px rgba(0,0,0,.06);padding:14px 16px;min-width:200px;flex:1;text-align:center}
.card h4{margin:0 0 6px;font-size:14px;color:#b02300;font-weight:700}
.card h2{margin:0;font-size:26px;font-weight:800;color:#222}
.card .sub{margin-top:8px;display:inline-block;padding:6px 10px;border-radius:8px;font-size:12px;font-weight:700}
.sub.ok{background:#e8f5ec;color:#197a31;border:1px solid #cce9d4}
.sub.bad{background:#fdeaea;color:#a31616;border:1px solid #f2cccc}
.sub.neu{background:#f1f1f4;color:#444;border:1px solid #e4e4e8}
.sub small{font-weight:600;color:#555;margin-left:8px}
.section{font-size:18px;font-weight:800;margin:22px 0 8px}
.small{color:#666;font-size:13px}
.table-note{margin-top:8px;color:#666;font-size:12px}
</style>
""",
    unsafe_allow_html=True,
)

# ⚡ MODO RÁPIDO (pula partes pesadas)
fast_mode = st.toggle("⚡ Modo rápido (carregar menos gráficos/tabelas pesadas)", value=False)


# ------------------ CREDENCIAL ------------------
def _get_client_and_drive():
    try:
        block = st.secrets["gcp_service_account"]
    except Exception:
        st.error("Não encontrei [gcp_service_account] no .streamlit/secrets.toml.")
        st.stop()

    if "json_path" in block:
        path = block["json_path"]
        if not os.path.isabs(path):
            path = os.path.join(os.path.dirname(__file__), path)
        try:
            with open(path, "r", encoding="utf-8") as f:
                info = json.load(f)
        except Exception as e:
            st.error(f"Não consegui abrir o JSON da service account: {path}")
            with st.expander("Detalhes"):
                st.exception(e)
            st.stop()
    else:
        info = dict(block)

    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(info, scopes)
    gc = gspread.authorize(creds)

    dscopes = ["https://www.googleapis.com/auth/drive.readonly"]
    gcred = gcreds.Credentials.from_service_account_info(info, scopes=dscopes)
    drive = build("drive", "v3", credentials=gcred, cache_discovery=False)

    return gc, drive, info.get("client_email", "(sem client_email)")


client, DRIVE, SA_EMAIL = _get_client_and_drive()


# ------------------ SECRETS: IDs ------------------
QUAL_INDEX_ID = st.secrets.get("qual_index_sheet_id", "").strip()
PROD_INDEX_ID = st.secrets.get("prod_index_sheet_id", "").strip()
if not QUAL_INDEX_ID:
    st.error("Faltou `qual_index_sheet_id` no secrets.toml"); st.stop()
if not PROD_INDEX_ID:
    st.error("Faltou `prod_index_sheet_id` no secrets.toml"); st.stop()


# ------------------ HELPERS ------------------
ID_RE = re.compile(r"/d/([a-zA-Z0-9-_]+)")

def _sheet_id(s: str) -> Optional[str]:
    s = (s or "").strip()
    m = ID_RE.search(s)
    if m:
        return m.group(1)
    return s if re.fullmatch(r"[A-Za-z0-9-_]{20,}", s) else None

def _ym_token(x: str) -> Optional[str]:
    """Converte 'MM/AAAA' -> 'AAAA-MM'."""
    if not x: return None
    s = str(x).strip()
    if re.fullmatch(r"\d{2}/\d{4}", s):
        mm, yy = s.split("/")
        return f"{yy}-{int(mm):02d}"
    if re.fullmatch(r"\d{4}-\d{2}", s):
        return s
    return None

def parse_date_any(x):
    if pd.isna(x) or x == "":
        return pd.NaT
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        try:
            return (pd.to_datetime("1899-12-30") + pd.to_timedelta(int(x), unit="D")).date()
        except Exception:
            pass
    s = str(x).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return pd.to_datetime(s).date()
    except Exception:
        return pd.NaT

def _upper(x):
    return str(x).upper().strip() if pd.notna(x) else ""

def _yes(v) -> bool:
    return str(v).strip().upper() in {"S", "SIM", "Y", "YES", "TRUE", "1"}

def _strip_accents(s: str) -> str:
    if s is None: return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(ch))

def _find_col(cols, *names) -> Optional[str]:
    """Encontra a coluna em 'cols' ignorando acentos/maiúsculas/espaços."""
    norm = {re.sub(r"\W+", "", _strip_accents(c).upper()): c for c in cols}
    for nm in names:
        key = re.sub(r"\W+", "", _strip_accents(nm).upper())
        if key in norm: return norm[key]
    return None

def business_days_count(dini: date, dfim: date) -> int:
    if not (isinstance(dini, date) and isinstance(dfim, date) and dini <= dfim):
        return 0
    return len(pd.bdate_range(dini, dfim))


# ------------------ LEITURA DOS ÍNDICES (com cache) ------------------
@st.cache_data(ttl=300, show_spinner=False)
def read_index(sheet_id: str, tab: str = "ARQUIVOS") -> pd.DataFrame:
    sh = client.open_by_key(sheet_id)
    ws = sh.worksheet(tab)
    rows = ws.get_all_records()
    if not rows:
        return pd.DataFrame(columns=["URL", "MÊS", "ATIVO"])
    df = pd.DataFrame(rows)
    df.columns = [c.strip().upper() for c in df.columns]
    for need in ["URL", "MÊS", "ATIVO"]:
        if need not in df.columns:
            df[need] = ""
    return df


# ------------------ FALLBACK XLSX / QUALIDADE (com cache) ------------------
@st.cache_data(ttl=300, show_spinner=False)
def _drive_get_file_metadata(file_id: str) -> dict:
    return DRIVE.files().get(fileId=file_id, fields="id, name, mimeType").execute()

@st.cache_data(ttl=300, show_spinner=False)
def _drive_download_bytes(file_id: str) -> bytes:
    req = DRIVE.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req, chunksize=1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()

@st.cache_data(ttl=300, show_spinner=False)
def read_quality_month(month_id: str) -> Tuple[pd.DataFrame, str]:
    meta = _drive_get_file_metadata(month_id)
    title = meta.get("name", month_id)
    mime = meta.get("mimeType", "")

    if mime == "application/vnd.google-apps.spreadsheet":
        sh = client.open_by_key(month_id)
        try:
            ws = sh.worksheet("GERAL")
        except Exception as e:
            raise RuntimeError(f"O arquivo '{title}' não possui aba 'GERAL'.") from e
        dq = pd.DataFrame(ws.get_all_records())
        if dq.empty:
            return pd.DataFrame(), title
        dq.columns = [c.strip() for c in dq.columns]
    else:
        if not mime.startswith("application/vnd.openxmlformats-officedocument") and \
           not mime.startswith("application/vnd.ms-excel"):
            raise RuntimeError(f"Tipo de arquivo não suportado para Qualidade: {mime} ({title})")
        content = _drive_download_bytes(month_id)
        try:
            dq = pd.read_excel(io.BytesIO(content), sheet_name="GERAL", engine="openpyxl")
        except ValueError as e:
            raise RuntimeError(f"O arquivo '{title}' não possui aba 'GERAL'.") from e
        dq.columns = [str(c).strip() for c in dq.columns]

    rename_map = {}
    for c in dq.columns:
        cu = c.upper()
        if cu == "DATA": rename_map[c] = "DATA"
        elif cu == "PLACA": rename_map[c] = "PLACA"
        elif cu in {"VISTORIADORES", "VISTORIADOR"}: rename_map[c] = "VISTORIADOR"
        elif cu in {"CIDADE", "UNIDADE"}: rename_map[c] = "UNIDADE"
        elif cu in {"ERROS","ERRO"}: rename_map[c] = "ERRO"
        elif cu.startswith("GRAVIDADE"): rename_map[c] = "GRAVIDADE"
        elif cu in {"OBSERVAÇÃO","OBSERVACAO","OBS"}: rename_map[c] = "OBS"
        elif cu == "ANALISTA": rename_map[c] = "ANALISTA"
        elif cu in {"EMPRESA","MARCA"}: rename_map[c] = "EMPRESA"
        elif cu in {"TEMPO DE CASA","TEMPO_DE_CASA","TEMPO CASA","TEMPOCASA"}:
            rename_map[c] = "TEMPO_CASA"
    dq = dq.rename(columns=rename_map)

    for need in ["DATA","PLACA","VISTORIADOR","UNIDADE","ERRO","GRAVIDADE","ANALISTA","EMPRESA","TEMPO_CASA"]:
        if need not in dq.columns:
            dq[need] = ""

    # Preserva timestamp e mantém DATA (date)
    if "DATA" in dq.columns:
        dq["DATA_TS"] = pd.to_datetime(dq["DATA"], errors="coerce")
        dq["DATA"] = dq["DATA"].apply(parse_date_any)
    else:
        dq["DATA_TS"] = pd.NaT

    for c in ["VISTORIADOR","UNIDADE","ERRO","GRAVIDADE","ANALISTA","EMPRESA","PLACA","TEMPO_CASA"]:
        dq[c] = dq[c].astype(str).map(_upper)

    dq = dq[(dq["VISTORIADOR"] != "") & (dq["ERRO"] != "")]
    return dq, title


# ------------------ LEITURA / PRODUÇÃO + METAS (com cache) ------------------
@st.cache_data(ttl=300, show_spinner=False)
def read_prod_month(month_sheet_id: str, ym: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame, str]:
    """Lê a planilha mensal de produção (aba 1) e, se existir, a aba 'METAS'."""
    sh = client.open_by_key(month_sheet_id)
    title = sh.title or month_sheet_id

    ws = sh.sheet1
    df = pd.DataFrame(ws.get_all_records())
    if not df.empty:
        df.columns = [c.strip().upper() for c in df.columns]

        col_unid = "UNIDADE" if "UNIDADE" in df.columns else None
        col_data = "DATA" if "DATA" in df.columns else None
        col_chas = "CHASSI" if "CHASSI" in df.columns else None
        col_per  = "PERITO" if "PERITO" in df.columns else None
        col_dig  = "DIGITADOR" if "DIGITADOR" in df.columns else None
        req = [col_unid, col_data, col_chas, (col_per or col_dig)]
        if any(r is None for r in req):
            df = pd.DataFrame()
        else:
            df[col_unid] = df[col_unid].map(_upper)
            df["__DATA__"] = df[col_data].apply(parse_date_any)
            df[col_chas] = df[col_chas].map(_upper)

            if col_per and col_dig:
                df["VISTORIADOR"] = np.where(
                    df[col_per].astype(str).str.strip() != "",
                    df[col_per].map(_upper),
                    df[col_dig].map(_upper),
                )
            elif col_per:
                df["VISTORIADOR"] = df[col_per].map(_upper)
            else:
                df["VISTORIADOR"] = df[col_dig].map(_upper)

            df = df.sort_values(["__DATA__", col_chas], kind="mergesort").reset_index(drop=True)
            df["__ORD__"] = df.groupby(col_chas).cumcount()
            df["IS_REV"] = (df["__ORD__"] >= 1).astype(int)
    metas = pd.DataFrame()

    try:
        ws_meta = sh.worksheet("METAS")
        rows = ws_meta.get_all_records()
        if rows:
            dm = pd.DataFrame(rows)
            cols = list(dm.columns)
            c_vist = _find_col(cols, "VISTORIADOR")
            c_unid = _find_col(cols, "UNIDADE")
            c_meta = _find_col(cols, "META_MENSAL", "META MENSAL", "META")
            c_du   = _find_col(cols, "DIAS ÚTEIS", "DIAS UTEIS", "DIAS_UTEIS")
            out = pd.DataFrame()
            out["VISTORIADOR"] = dm[c_vist].astype(str).map(_upper) if c_vist else ""
            out["UNIDADE"] = dm[c_unid].astype(str).map(_upper) if c_unid else ""
            out["META_MENSAL"] = pd.to_numeric(dm[c_meta], errors="coerce").fillna(0).astype(int) if c_meta else 0
            out["DIAS_UTEIS"]  = pd.to_numeric(dm[c_du], errors="coerce").fillna(np.nan)
            out["DIAS_UTEIS"]  = out["DIAS_UTEIS"].astype(float).round().astype("Int64")
            out["YM"] = ym or ""
            metas = out
    except Exception:
        metas = pd.DataFrame()

    return df, metas, title


# ------------------ CARREGA INDEX ------------------
show_tech = False

idx_q = read_index(QUAL_INDEX_ID)
if "ATIVO" in idx_q.columns:
    idx_q = idx_q[idx_q["ATIVO"].map(_yes)].copy()
sel_meses = sorted([str(m).strip() for m in idx_q["MÊS"] if str(m).strip()])

idx_p = read_index(PROD_INDEX_ID)
if "ATIVO" in idx_p.columns:
    idx_p = idx_p[idx_p["ATIVO"].map(_yes)].copy()
sel_meses_p = sorted([str(m).strip() for m in idx_p["MÊS"] if str(m).strip()])

if sel_meses:
    idx_q = idx_q[idx_q["MÊS"].isin(sel_meses)]
if sel_meses_p:
    idx_p = idx_p[idx_p["MÊS"].isin(sel_meses_p)]

dq_all, ok_q, er_q = [], [], []
for _, r in idx_q.iterrows():
    sid = _sheet_id(r["URL"])
    if not sid: continue
    try:
        dq, ttl = read_quality_month(sid)
        if not dq.empty: dq_all.append(dq)
        ok_q.append(f"✅ {ttl} — {len(dq):,} linhas".replace(",", "."))
    except Exception as e:
        er_q.append((sid, e))

dp_all, metas_all, ok_p, er_p = [], [], [], []
for _, r in idx_p.iterrows():
    sid = _sheet_id(r["URL"])
    ym  = _ym_token(r.get("MÊS", ""))
    if not sid: continue
    try:
        dp, dm, ttl = read_prod_month(sid, ym=ym)
        if not dp.empty:    dp_all.append(dp)
        if not dm.empty:    metas_all.append(dm)
        ok_p.append(f"✅ {ttl} — {len(dp):,} linhas")
    except Exception as e:
        er_p.append((sid, e))

if show_tech:
    if ok_q: st.success("Qualidade conectado em:\n\n- " + "\n- ".join(ok_q))
    if er_q:
        with st.expander("Falhas (Qualidade)"):
            for sid, e in er_q: st.write(sid); st.exception(e)
    if ok_p: st.success("Produção conectado em:\n\n- " + "\n- ".join(ok_p))
    if er_p:
        with st.expander("Falhas (Produção)"):
            for sid, e in er_p: st.write(sid); st.exception(e)

if not dq_all:
    st.error("Não consegui ler dados de Qualidade de nenhum mês."); st.stop()

dfQ = pd.concat(dq_all, ignore_index=True)
dfP = pd.concat(dp_all, ignore_index=True) if dp_all else pd.DataFrame(columns=["VISTORIADOR","__DATA__","IS_REV","UNIDADE"])
dfMetas = pd.concat(metas_all, ignore_index=True) if metas_all else pd.DataFrame(columns=["VISTORIADOR","UNIDADE","META_MENSAL","DIAS_UTEIS","YM"])

# Normaliza TEMPO_CASA (NOVATO / VETERANO) — IGUAL TOKYO
if "TEMPO_CASA" in dfQ.columns:
    dfQ["TEMPO_CASA"] = dfQ["TEMPO_CASA"].astype(str).map(_upper)


# ------------------ FILTROS PRINCIPAIS ------------------
if "EMPRESA" in dfQ.columns:
    dfQ = dfQ[dfQ["EMPRESA"] == "VELOX"].copy()

s_all_dt = pd.to_datetime(dfQ["DATA"], errors="coerce")
ym_all = sorted(s_all_dt.dt.to_period("M").dropna().astype(str).unique().tolist())
if not ym_all:
    st.error("Qualidade sem colunas de Data válidas."); st.stop()

label_map = {f"{m[5:]}/{m[:4]}": m for m in ym_all}
sel_label = st.selectbox("Mês de referência", options=list(label_map.keys()), index=len(ym_all)-1)
ym_sel = label_map[sel_label]
ref_year, ref_month = int(ym_sel[:4]), int(ym_sel[5:7])

mask_mes = (s_all_dt.dt.year.eq(ref_year) & s_all_dt.dt.month.eq(ref_month))
dfQ_mes = dfQ[mask_mes].copy()

s_mes_dates = pd.to_datetime(dfQ_mes["DATA"], errors="coerce").dt.date
min_d, max_d = min(s_mes_dates.dropna()), max(s_mes_dates.dropna())
col1, col2 = st.columns([1.2, 2.8])
with col1:
    drange = st.date_input(
        "Período (dentro do mês)",
        value=(min_d, max_d), min_value=min_d, max_value=max_d,
        format="DD/MM/YYYY"
    )

start_d, end_d = (drange if isinstance(drange, tuple) and len(drange)==2 else (min_d, max_d))
mask_dias = s_mes_dates.map(lambda d: isinstance(d, date) and start_d <= d <= end_d)
viewQ = dfQ_mes[mask_dias].copy()

# -------- Filtros extras --------
unids = sorted(viewQ["UNIDADE"].dropna().unique().tolist()) if "UNIDADE" in viewQ.columns else []
vist_opts = sorted(viewQ["VISTORIADOR"].dropna().unique().tolist()) if "VISTORIADOR" in viewQ.columns else []

with col2:
    c21, c22 = st.columns(2)
    with c21:
        f_unids = st.multiselect("Unidades (opcional)", unids, default=unids)
    with c22:
        f_vists = st.multiselect("Vistoriadores (opcional)", vist_opts)
    # Filtro de tempo de casa no cabeçalho — IGUAL TOKYO
    perfil_sel = st.radio(
        "Tempo de casa",
        ["Todos", "Somente Novatos", "Somente Veteranos"],
        horizontal=True,
        key="perfil_tempo_casa_velox"
    )

# Ajusta label para NOVATO / VETERANO interno
if perfil_sel == "Somente Novatos":
    alvo_perfil = "NOVATO"
elif perfil_sel == "Somente Veteranos":
    alvo_perfil = "VETERANO"
else:
    alvo_perfil = None

if f_unids and "UNIDADE" in viewQ.columns:
    viewQ = viewQ[viewQ["UNIDADE"].isin([_upper(u) for u in f_unids])]
if f_vists:
    viewQ = viewQ[viewQ["VISTORIADOR"].isin([_upper(v) for v in f_vists])]

# Aplica filtro NOVATO / VETERANO na base de qualidade
set_vists_perfil = None
if "TEMPO_CASA" in viewQ.columns and alvo_perfil is not None:
    viewQ = viewQ[viewQ["TEMPO_CASA"] == alvo_perfil]
    set_vists_perfil = set(viewQ["VISTORIADOR"].unique())

if viewQ.empty:
    st.info("Sem registros de Qualidade no período/filtros."); st.stop()

# -------- Produção alinhada --------
if not dfP.empty:
    s_p_dates_all = pd.to_datetime(dfP["__DATA__"], errors="coerce").dt.date
    maskp_mes = s_p_dates_all.map(lambda d: isinstance(d, date) and d.year == ref_year and d.month == ref_month)
    viewP = dfP[maskp_mes].copy()

    s_p_dates_mes = pd.to_datetime(viewP["__DATA__"], errors="coerce").dt.date
    maskp_dias = s_p_dates_mes.map(lambda d: isinstance(d, date) and start_d <= d <= end_d)
    viewP = viewP[maskp_dias].copy()

    if f_unids and "UNIDADE" in viewP.columns:
        viewP = viewP[viewP["UNIDADE"].isin([_upper(u) for u in f_unids])]
    if f_vists and "VISTORIADOR" in viewP.columns:
        viewP = viewP[viewP["VISTORIADOR"].isin([_upper(v) for v in f_vists])]

    # Aplica filtro de perfil também na produção
    if set_vists_perfil is not None and "VISTORIADOR" in viewP.columns:
        viewP = viewP[viewP["VISTORIADOR"].isin(set_vists_perfil)]
else:
    viewP = dfP.copy()
    if set_vists_perfil is not None and "VISTORIADOR" in viewP.columns:
        viewP = viewP[viewP["VISTORIADOR"].isin(set_vists_perfil)]


# ------------------ KPIs ------------------
grav_gg = {"GRAVE", "GRAVISSIMO", "GRAVÍSSIMO"}
total_erros = int(len(viewQ))
total_gg = int(viewQ["GRAVIDADE"].isin(grav_gg).sum()) if "GRAVIDADE" in viewQ.columns else 0
vist_avaliados = int(viewQ["VISTORIADOR"].nunique()) if "VISTORIADOR" in viewQ.columns else 0
media_por_vist = (total_erros / vist_avaliados) if vist_avaliados else 0

if "GRAVIDADE" in viewQ.columns:
    gg_by_vist = (
        viewQ[viewQ["GRAVIDADE"].isin(grav_gg)]
        .groupby("VISTORIADOR")["ERRO"].size().reset_index(name="GG")
    )
    vist_5gg = int((gg_by_vist["GG"] >= 5).sum())
else:
    vist_5gg = 0

total_vist_brutas = int(len(viewP)) if not viewP.empty else 0
taxa_geral = (total_erros / total_vist_brutas * 100) if total_vist_brutas else np.nan
taxa_geral_str = "—" if np.isnan(taxa_geral) else f"{taxa_geral:.1f}%".replace(".", ",")

# % GG sobre produção bruta
taxa_gg_bruta = (total_gg / total_vist_brutas * 100) if total_vist_brutas else np.nan
taxa_gg_bruta_str = "—" if np.isnan(taxa_gg_bruta) else f"{taxa_gg_bruta:.1f}%".replace(".", ",")

# ---- Comparativo com mesmo intervalo do mês anterior (para os cards) ----
periodo_atual_ini, periodo_atual_fim = start_d, end_d
prev_ini = (pd.Timestamp(periodo_atual_ini) - relativedelta(months=1)).date()
prev_fim = (pd.Timestamp(periodo_atual_fim) - relativedelta(months=1)).date()

dfQ["_DT_"] = pd.to_datetime(dfQ["DATA"], errors="coerce").dt.date
mask_prev = dfQ["_DT_"].between(prev_ini, prev_fim)
prev_base_cards = dfQ[mask_prev].copy()
if "UNIDADE" in prev_base_cards.columns and len(f_unids):
    prev_base_cards = prev_base_cards[prev_base_cards["UNIDADE"].isin([_upper(u) for u in f_unids])]
if "VISTORIADOR" in prev_base_cards.columns and len(f_vists):
    prev_base_cards = prev_base_cards[prev_base_cards["VISTORIADOR"].isin([_upper(v) for v in f_vists])]
if "TEMPO_CASA" in prev_base_cards.columns and alvo_perfil is not None:
    prev_base_cards = prev_base_cards[prev_base_cards["TEMPO_CASA"] == alvo_perfil]

prev_total = int(len(prev_base_cards))
prev_gg = int(prev_base_cards["GRAVIDADE"].isin(grav_gg).sum()) if "GRAVIDADE" in prev_base_cards.columns else 0

def _pct_delta(cur, prev):
    if prev <= 0:
        return None
    return (cur - prev) / prev * 100.0

def _badge_html(delta_pct, prev_value):
    if delta_pct is None:
        cls, txt = "neu", "—"
    else:
        sign = "+" if delta_pct >= 0 else ""
        status = "Piorou" if delta_pct > 0 else ("Melhorou" if delta_pct < 0 else "Igual")
        cls = "bad" if delta_pct > 0 else ("ok" if delta_pct < 0 else "neu")
        txt = f"{sign}{delta_pct:.1f}% · {status}"
    prev_txt = f"<small>mês ant: {prev_value:,}</small>".replace(",", ".")
    return f"<span class='sub {cls}'>{txt} {prev_txt}</span>"

delta_total = _pct_delta(total_erros, prev_total)
delta_gg    = _pct_delta(total_gg, prev_gg)
badge_total = _badge_html(delta_total, prev_total)
badge_gg    = _badge_html(delta_gg, prev_gg)

# ---- Projeções do mês (marca) ----
month_start = date(ref_year, ref_month, 1)
last_day = calendar.monthrange(ref_year, ref_month)[1]
month_end = date(ref_year, ref_month, last_day)

dfQ["_DTONLY_"] = pd.to_datetime(dfQ["DATA"], errors="coerce").dt.date
mask_mtd = dfQ["_DTONLY_"].between(month_start, min(end_d, month_end))
mtd_all = dfQ[mask_mtd].copy()
if "UNIDADE" in mtd_all.columns and len(f_unids):
    mtd_all = mtd_all[mtd_all["UNIDADE"].isin([_upper(u) for u in f_unids])]
if "VISTORIADOR" in mtd_all.columns and len(f_vists):
    mtd_all = mtd_all[mtd_all["VISTORIADOR"].isin([_upper(v) for v in f_vists])]
if "TEMPO_CASA" in mtd_all.columns and alvo_perfil is not None:
    mtd_all = mtd_all[mtd_all["TEMPO_CASA"] == alvo_perfil]

erros_mtd_total = int(len(mtd_all))
erros_mtd_gg = int(mtd_all["GRAVIDADE"].isin(grav_gg).sum()) if "GRAVIDADE" in mtd_all.columns else 0

dias_passados = business_days_count(month_start, min(end_d, month_end))
dias_totais_fallback = business_days_count(month_start, month_end)

def _proj(cur_mtd):
    if dias_passados == 0:
        return cur_mtd
    return int(round(cur_mtd / dias_passados * dias_totais_fallback))

proj_total = _proj(erros_mtd_total)
proj_gg = _proj(erros_mtd_gg)

# ------------------ CARDS ------------------
cards_html = """
<div class="card-wrap">
  <div class='card'>
    <h4>Total de erros (período)</h4>
    <h2>{total_erros}</h2>
    {badge_total}
  </div>
  <div class='card'>
    <h4>Vistoriadores com ≥5 erros GG</h4>
    <h2>{vist_5gg}</h2>
  </div>
  <div class='card'>
    <h4>Erros Grave+Gravíssimo</h4>
    <h2>{total_gg}</h2>
    {badge_gg}
  </div>
  <div class='card'>
    <h4>Vistoriadores avaliados</h4>
    <h2>{vist_avaliados}</h2>
  </div>
  <div class='card'>
    <h4>Média de erros / vistoriador</h4>
    <h2>{media_por_vist}</h2>
  </div>
  <div class='card'>
    <h4>Taxa de erro (bruta)</h4>
    <h2>{taxa_geral}</h2>
  </div>
  <div class='card'>
    <h4>% GG sobre a produção</h4>
    <h2>{taxa_gg_bruta}</h2>
    <span class='sub neu'>base: vistorias brutas</span>
  </div>
  <div class='card'>
    <h4>Projeção do mês — Erros</h4>
    <h2>{proj_total}</h2>
    <span class='sub neu'>MTD: {mtd_total}</span>
  </div>
  <div class='card'>
    <h4>Projeção do mês — Erros GG</h4>
    <h2>{proj_gg}</h2>
    <span class='sub neu'>MTD: {mtd_gg}</span>
  </div>
</div>
""".format(
    total_erros=f"{total_erros:,}".replace(",", "."),
    badge_total=badge_total,
    vist_5gg=f"{vist_5gg:,}".replace(",", "."),
    total_gg=f"{total_gg:,}".replace(",", "."),
    badge_gg=badge_gg,
    vist_avaliados=f"{vist_avaliados:,}".replace(",", "."),
    media_por_vist=f"{media_por_vist:.1f}".replace(".", ","),
    taxa_geral=taxa_geral_str,
    taxa_gg_bruta=taxa_gg_bruta_str,
    proj_total=f"{proj_total:,}".replace(",", "."),
    proj_gg=f"{proj_gg:,}".replace(",", "."),
    mtd_total=f"{erros_mtd_total:,}".replace(",", "."),
    mtd_gg=f"{erros_mtd_gg:,}".replace(",", "."),
)
st.markdown(cards_html, unsafe_allow_html=True)

# ------------------ BASE BRUTA x LÍQUIDA (global p/ heatmap e %Erro) ------------------
denom_mode = st.radio(
    "Base para %Erro (usada no heatmap e na tabela de % por vistoriador)",
    ["Bruta (recomendado)", "Líquida"],
    horizontal=True, index=0, key="denom_mode_global"
)

# ------------------ HOJE x ONTEM (ATÉ AGORA) ------------------
st.markdown('<div class="section">⏱️ Hoje vs Ontem (até agora)</div>', unsafe_allow_html=True)

try:
    from zoneinfo import ZoneInfo
    tz = ZoneInfo("America/Fortaleza")
except Exception:
    tz = None

now_local = datetime.now(tz) if tz else datetime.now()
today_local = now_local.date()
yesterday_local = (now_local - pd.Timedelta(days=1)).date()

def _as_naive_ts(series_like):
    ts = pd.to_datetime(series_like, errors="coerce")
    try:
        if getattr(ts.dt, "tz", None) is not None:
            try:
                ts = ts.dt.tz_convert(None)
            except Exception:
                ts = ts.dt.tz_localize(None)
    except Exception:
        pass
    return ts

def _as_naive_cutoff(dt_like):
    ts = pd.Timestamp(dt_like).replace(second=0, microsecond=0)
    if ts.tz is not None:
        ts = ts.tz_localize(None)
    return ts

if start_d == end_d == today_local:
    df_today = viewQ.copy()
    if "DATA_TS" not in df_today.columns:
        df_today["DATA_TS"] = pd.to_datetime(df_today["DATA"], errors="coerce")

    ts_today = _as_naive_ts(df_today["DATA_TS"])
    have_time_today = ts_today.dt.hour.notna().any()

    cutoff_today = _as_naive_cutoff(now_local)
    if have_time_today:
        mask_today_now = (ts_today <= cutoff_today)
        df_today_now = df_today[mask_today_now]
    else:
        df_today_now = df_today

    df_all = dfQ.copy()
    mask_yesterday = pd.to_datetime(df_all["DATA"], errors="coerce").dt.date.eq(yesterday_local)
    df_yest = df_all[mask_yesterday].copy()
    if len(f_unids) and "UNIDADE" in df_yest.columns:
        df_yest = df_yest[df_yest["UNIDADE"].isin([_upper(u) for u in f_unids])]
    if len(f_vists) and "VISTORIADOR" in df_yest.columns:
        df_yest = df_yest[df_yest["VISTORIADOR"].isin([_upper(v) for v in f_vists])]
    if "TEMPO_CASA" in df_yest.columns and alvo_perfil is not None:
        df_yest = df_yest[df_yest["TEMPO_CASA"] == alvo_perfil]

    if "DATA_TS" not in df_yest.columns:
        df_yest["DATA_TS"] = pd.to_datetime(df_yest["DATA"], errors="coerce")

    ts_yest = _as_naive_ts(df_yest["DATA_TS"])
    have_time_yest = ts_yest.dt.hour.notna().any()

    if have_time_today and have_time_yest:
        cutoff_yest = _as_naive_cutoff(
            now_local.replace(year=yesterday_local.year, month=yesterday_local.month, day=yesterday_local.day)
        )
        mask_yest_now = (ts_yest <= cutoff_yest)
        df_yest_now = df_yest[mask_yest_now]
        note_text = "Comparando até a mesma hora (base com horário)."
    else:
        df_yest_now = df_yest
        note_text = "Sem horário na base — comparando o dia inteiro."

    erros_hoje_ate_agora = int(len(df_today_now))
    erros_ontem_mesma_hora = int(len(df_yest_now))
    delta = erros_hoje_ate_agora - erros_ontem_mesma_hora
    tendencia = "❌ Piorou" if delta > 0 else ("✅ Melhorou" if delta < 0 else "➡️ Igual")

    cA, cB, cC = st.columns([1, 1, 1])
    cA.metric("Erros HOJE (até agora)", f"{erros_hoje_ate_agora:,}".replace(",", "."), delta=f"{delta:+d} vs ontem")
    cB.metric("Erros ONTEM (mesma hora)", f"{erros_ontem_mesma_hora:,}".replace(",", "."))
    cC.metric("Tendência", tendencia)

    st.caption(f"<span class='small'>{note_text}</span>", unsafe_allow_html=True)
else:
    st.info("Para ver o comparativo HOJE x ONTEM, selecione o dia atual no filtro de período.")


# ------------------ GRÁFICOS ------------------
def bar_with_labels(df, x_col, y_col, x_title="", y_title="QTD", height=320):
    base = alt.Chart(df).encode(
        x=alt.X(f"{x_col}:N", sort='-y', title=x_title,
                axis=alt.Axis(labelAngle=0, labelLimit=180, labelOverlap=False)),
        y=alt.Y(f"{y_col}:Q", title=y_title),
        tooltip=[x_col, y_col],
    )
    bars = base.mark_bar()
    labels = base.mark_text(dy=-6).encode(text=alt.Text(f"{y_col}:Q", format=".0f"))
    return (bars + labels).properties(height=height)

c1, c2 = st.columns(2)

if "UNIDADE" in viewQ.columns:
    with c1:
        st.markdown('<div class="section">🏙️ Erros por unidade</div>', unsafe_allow_html=True)

        # duas colunas: TOTAL (à esquerda) e GG (à direita)
        g_tot, g_gg = st.columns(2)

        # ---------- TOTAL de erros por unidade ----------
        with g_tot:
            by_city = (
                viewQ.groupby("UNIDADE", dropna=False)["ERRO"].size().reset_index(name="QTD")
            )

            if not viewP.empty and "UNIDADE" in viewP.columns:
                prod_city = (
                    viewP.groupby("UNIDADE", dropna=False)["IS_REV"].size().reset_index(name="VIST")
                )
            else:
                prod_city = pd.DataFrame(columns=["UNIDADE", "VIST"])

            by_city = by_city.merge(prod_city, on="UNIDADE", how="left").fillna({"VIST": 0})
            by_city["%ERRO"] = np.where(by_city["VIST"] > 0, (by_city["QTD"] / by_city["VIST"]) * 100, np.nan)

            if by_city["%ERRO"].isna().all():
                total_err = by_city["QTD"].sum()
                by_city["%ERRO"] = np.where(total_err > 0, (by_city["QTD"] / total_err) * 100, np.nan)
                y2_title = "% dos erros"
            else:
                y2_title = "% de erro (erros/vistorias)"

            by_city["PCT"] = by_city["%ERRO"] / 100.0
            by_city = by_city.sort_values("QTD", ascending=False).reset_index(drop=True)
            order = by_city["UNIDADE"].tolist()

            bars = (
                alt.Chart(by_city).mark_bar().encode(
                    x=alt.X("UNIDADE:N", sort=order, axis=alt.Axis(labelAngle=0, labelLimit=180), title="UNIDADE"),
                    y=alt.Y("QTD:Q", title="QTD"),
                    tooltip=["UNIDADE", "QTD", alt.Tooltip("PCT:Q", format=".1%", title=y2_title)],
                )
            )
            bar_labels = (
                alt.Chart(by_city).mark_text(dy=-6).encode(
                    x=alt.X("UNIDADE:N", sort=order),
                    y="QTD:Q",
                    text=alt.Text("QTD:Q", format=".0f"),
                )
            )
            line = (
                alt.Chart(by_city).mark_line(point=True, color="#b02300").encode(
                    x=alt.X("UNIDADE:N", sort=order),
                    y=alt.Y("PCT:Q", axis=alt.Axis(title=y2_title, format=".1%")),
                )
            )
            line_labels = (
                alt.Chart(by_city).mark_text(color="#b02300", dy=-8, fontWeight="bold").encode(
                    x=alt.X("UNIDADE:N", sort=order),
                    y="PCT:Q",
                    text=alt.Text("PCT:Q", format=".1%"),
                )
            )
            chart = alt.layer(bars, bar_labels, line, line_labels).resolve_scale(y="independent").properties(height=340)
            st.subheader("Total")
            st.altair_chart(chart, use_container_width=True)

        # ---------- Somente GRAVE + GRAVÍSSIMO por unidade ----------
        with g_gg:
            mask_gg = viewQ["GRAVIDADE"].astype(str).str.upper().isin(grav_gg) if "GRAVIDADE" in viewQ.columns else pd.Series(False, index=viewQ.index)
            viewQ_gg = viewQ[mask_gg]

            by_city_gg = (
                viewQ_gg.groupby("UNIDADE", dropna=False)["ERRO"].size().reset_index(name="QTD_GG")
            )

            if not viewP.empty and "UNIDADE" in viewP.columns:
                prod_city = (
                    viewP.groupby("UNIDADE", dropna=False)["IS_REV"].size().reset_index(name="VIST")
                )
            else:
                prod_city = pd.DataFrame(columns=["UNIDADE", "VIST"])

            by_city_gg = by_city_gg.merge(prod_city, on="UNIDADE", how="left").fillna({"VIST": 0})

            by_city_gg["%ERRO_GG"] = np.where(by_city_gg["VIST"] > 0,
                                              (by_city_gg["QTD_GG"] / by_city_gg["VIST"]) * 100, np.nan)
            if by_city_gg["%ERRO_GG"].isna().all():
                total_gg_global = by_city_gg["QTD_GG"].sum()
                by_city_gg["%ERRO_GG"] = np.where(total_gg_global > 0,
                                                  (by_city_gg["QTD_GG"] / total_gg_global) * 100, np.nan)
                y2_title_gg = "% dos erros GG"
            else:
                y2_title_gg = "% de erro GG (GG/vistorias)"

            by_city_gg["PCT_GG"] = by_city_gg["%ERRO_GG"] / 100.0
            by_city_gg = by_city_gg.sort_values("QTD_GG", ascending=False).reset_index(drop=True)
            order_gg = by_city_gg["UNIDADE"].tolist()

            bars_gg = (
                alt.Chart(by_city_gg).mark_bar().encode(
                    x=alt.X("UNIDADE:N", sort=order_gg, axis=alt.Axis(labelAngle=0, labelLimit=180), title="UNIDADE"),
                    y=alt.Y("QTD_GG:Q", title="QTD (GG)"),
                    tooltip=["UNIDADE", "QTD_GG", alt.Tooltip("PCT_GG:Q", format=".1%", title=y2_title_gg)],
                )
            )
            bar_labels_gg = (
                alt.Chart(by_city_gg).mark_text(dy=-6).encode(
                    x=alt.X("UNIDADE:N", sort=order_gg),
                    y="QTD_GG:Q",
                    text=alt.Text("QTD_GG:Q", format=".0f"),
                )
            )
            line_gg = (
                alt.Chart(by_city_gg).mark_line(point=True, color="#b02300").encode(
                    x=alt.X("UNIDADE:N", sort=order_gg),
                    y=alt.Y("PCT_GG:Q", axis=alt.Axis(title=y2_title_gg, format=".1%")),
                )
            )
            line_labels_gg = (
                alt.Chart(by_city_gg).mark_text(color="#b02300", dy=-8, fontWeight="bold").encode(
                    x=alt.X("UNIDADE:N", sort=order_gg),
                    y="PCT_GG:Q",
                    text=alt.Text("PCT_GG:Q", format=".1%"),
                )
            )
            chart_gg = alt.layer(bars_gg, bar_labels_gg, line_gg, line_labels_gg).resolve_scale(y="independent").properties(height=340)
            st.subheader("Grave + Gravíssimo")
            st.altair_chart(chart_gg, use_container_width=True)

if "GRAVIDADE" in viewQ.columns:
    with c2:
        st.markdown('<div class="section">🧲 Erros por gravidade</div>', unsafe_allow_html=True)
        by_grav = (viewQ.groupby("GRAVIDADE", dropna=False)["ERRO"]
                   .size().reset_index(name="QTD").sort_values("QTD", ascending=False))
        if len(by_grav):
            st.altair_chart(bar_with_labels(by_grav, "GRAVIDADE", "QTD", x_title="GRAVIDADE", height=340),
                            use_container_width=True)

# ------------------ TOP 5 ERROS GRAVES / GRAVÍSSIMOS ------------------
st.markdown("---")
st.markdown('<div class="section">🥇 Top 5 — erros GRAVE e GRAVÍSSIMO</div>', unsafe_allow_html=True)

if "GRAVIDADE" in viewQ.columns:
    grav_alias_gravissimo = {"GRAVISSIMO", "GRAVÍSSIMO"}

    mask_grave = viewQ["GRAVIDADE"].astype(str).str.upper().eq("GRAVE")
    mask_gravissimo = viewQ["GRAVIDADE"].astype(str).str.upper().isin(grav_alias_gravissimo)

    top_grave = (
        viewQ[mask_grave]
        .groupby("ERRO", dropna=False)["ERRO"].size()
        .reset_index(name="QTD")
        .sort_values("QTD", ascending=False)
        .head(5)
    )

    top_gravissimo = (
        viewQ[mask_gravissimo]
        .groupby("ERRO", dropna=False)["ERRO"].size()
        .reset_index(name="QTD")
        .sort_values("QTD", ascending=False)
        .head(5)
    )

    cG, cGG = st.columns(2)
    with cG:
        st.subheader("Top 5 — GRAVE")
        if top_grave.empty:
            st.info("Sem erros GRAVE no recorte atual.")
        else:
            st.altair_chart(
                bar_with_labels(top_grave, "ERRO", "QTD", x_title="ERRO (GRAVE)", y_title="QTD", height=320),
                use_container_width=True
            )

    with cGG:
        st.subheader("Top 5 — GRAVÍSSIMO")
        if top_gravissimo.empty:
            st.info("Sem erros GRAVÍSSIMO no recorte atual.")
        else:
            st.altair_chart(
                bar_with_labels(top_gravissimo, "ERRO", "QTD", x_title="ERRO (GRAVÍSSIMO)", y_title="QTD", height=320),
                use_container_width=True
            )
else:
    st.info("Base sem coluna de GRAVIDADE para montar os Top 5.")

# ------------------ VISUALIZAÇÕES EXTRAS ------------------
if not fast_mode:
    ex1, ex2 = st.columns(2)

    # ===== PARETO =====
    with ex1:
        st.markdown('<div class="section">📈 Pareto de erros</div>', unsafe_allow_html=True)

        n_err = int(viewQ["ERRO"].nunique()) if "ERRO" in viewQ.columns else 0
        if n_err == 0:
            st.info("Sem dados para montar o Pareto no período/filtros atuais.")
        else:
            max_cats = min(30, n_err)
            if max_cats < 1:
                st.info("Sem categorias suficientes para montar o Pareto.")
            else:
                if max_cats <= 1:
                    top_cats = 1
                    st.caption("Categorias no Pareto: 1")
                else:
                    top_default = min(10, max_cats)
                    top_cats = st.slider(
                        "Categorias no Pareto",
                        min_value=1, max_value=max_cats, value=top_default,
                        step=1, key=f"pareto_cats_{ref_year}{ref_month}",
                    )

                pareto = (
                    viewQ.groupby("ERRO", sort=False)["ERRO"]
                    .size()
                    .reset_index(name="QTD")
                    .sort_values("QTD", ascending=False)
                    .head(top_cats)
                    .reset_index(drop=True)
                )

                if pareto.empty:
                    st.info("Sem dados para montar o Pareto no período/filtros atuais.")
                else:
                    pareto["ACUM"] = pareto["QTD"].cumsum()
                    total = pareto["QTD"].sum()
                    pareto["%ACUM"] = pareto["ACUM"] / total * 100

                    x_enc = alt.X(
                        "ERRO:N",
                        sort=alt.SortField(field="QTD", order="descending"),
                        axis=alt.Axis(labelAngle=0, labelLimit=180),
                        title="ERRO",
                    )

                    bars = alt.Chart(pareto).mark_bar().encode(
                        x=x_enc,
                        y=alt.Y("QTD:Q", title="QTD"),
                        tooltip=["ERRO", "QTD", alt.Tooltip("%ACUM:Q", format=".1f", title="% acumulado")],
                    )
                    bar_labels = alt.Chart(pareto).mark_text(dy=-6).encode(
                        x=x_enc, y="QTD:Q", text=alt.Text("QTD:Q", format=".0f")
                    )

                    line = alt.Chart(pareto).mark_line(point=True).encode(
                        x=x_enc,
                        y=alt.Y("%ACUM:Q", title="% Acumulado"),
                        color=alt.value("#b02300"),
                    )
                    line_labels = (
                        alt.Chart(pareto)
                        .mark_text(dy=-8, baseline="bottom", color="#b02300", fontWeight="bold")
                        .encode(x=x_enc, y="%ACUM:Q", text=alt.Text("%ACUM:Q", format=".1f"))
                    )

                    chart_pareto = (
                        alt.layer(bars, bar_labels, line, line_labels)
                        .resolve_scale(y="independent")
                        .properties(height=360)
                    )
                    st.altair_chart(chart_pareto, use_container_width=True)

                    max_topN = int(len(pareto))
                    if max_topN <= 1:
                        topN_sim = 1
                        st.caption("Top considerado: 1")
                        reducao = st.slider(
                            "Redução esperada nesses erros (%)",
                            min_value=0, max_value=100, value=25,
                            key=f"pareto_reducao_{ref_year}{ref_month}",
                        )
                    else:
                        topN_sim = st.slider(
                            "Quantos erros do topo considerar?",
                            min_value=1, max_value=max_topN, value=min(8, max_topN),
                            key=f"pareto_topN_{ref_year}{ref_month}",
                        )
                        reducao = st.slider(
                            "Redução esperada nesses erros (%)",
                            min_value=0, max_value=100, value=25,
                            key=f"pareto_reducao_{ref_year}{ref_month}",
                        )

                    idx = min(topN_sim, max_topN) - 1
                    frac = float(pareto["%ACUM"].iloc[idx]) / 100.0
                    queda_total = frac * (reducao / 100.0) * 100.0

                    st.info(
                        f"Os Top {topN_sim} explicam {frac*100:.1f}% do total. "
                        f"Se reduzir esses erros em {reducao}%, o total cai cerca de {queda_total:.1f}%."
                    )

    with ex2:
        st.markdown('<div class="section">🗺️ Heatmap Cidade × Gravidade</div>', unsafe_allow_html=True)
        if ("UNIDADE" in viewQ.columns) and ("GRAVIDADE" in viewQ.columns):
            # Erros por UNIDADE x GRAVIDADE
            erros_city = (
                viewQ.groupby(["UNIDADE", "GRAVIDADE"])["ERRO"]
                .size()
                .reset_index(name="QTD")
            )

            # Denominador: vistorias por cidade no mesmo recorte (Bruta/Líquida conforme rádio)
            if not viewP.empty and "UNIDADE" in viewP.columns:
                prod_city = (
                    viewP.groupby("UNIDADE", dropna=False)
                    .agg(vist=("IS_REV", "size"), rev=("IS_REV", "sum"))
                    .reset_index()
                )
                prod_city["liq"] = prod_city["vist"] - prod_city["rev"]
            else:
                prod_city = pd.DataFrame({
                    "UNIDADE": erros_city["UNIDADE"].unique(),
                    "vist": 0, "rev": 0
                })
                prod_city["liq"] = 0

            denom_col = "liq" if denom_mode.startswith("Líquida") else "vist"

            hm = erros_city.merge(
                prod_city[["UNIDADE", denom_col]].rename(columns={denom_col: "DEN"}),
                on="UNIDADE",
                how="left",
            )
            hm["%_VIST"] = np.where(hm["DEN"] > 0, (hm["QTD"] / hm["DEN"]) * 100, np.nan)
            hm["%_VIST_TXT"] = hm["%_VIST"].map(lambda x: "—" if pd.isna(x) else f"{x:.1f}%".replace(".", ","))

            rects = alt.Chart(hm).mark_rect().encode(
                x=alt.X("GRAVIDADE:N", axis=alt.Axis(labelAngle=0, title="GRAVIDADE")),
                y=alt.Y("UNIDADE:N", sort='-x', title="UNIDADE"),
                color=alt.Color("QTD:Q", scale=alt.Scale(scheme="blues"), title="QTD"),
                tooltip=[
                    alt.Tooltip("UNIDADE:N", title="UNIDADE"),
                    alt.Tooltip("GRAVIDADE:N", title="GRAVIDADE"),
                    alt.Tooltip("QTD:Q", format=".0f", title="Erros"),
                    alt.Tooltip("DEN:Q", format=".0f",
                                title=f"Vistorias ({'líq.' if denom_col=='liq' else 'brutas'})"),
                    alt.Tooltip("%_VIST_TXT:N", title="% sobre vistorias"),
                ],
            )

            labels = alt.Chart(hm).mark_text(baseline="middle").encode(
                x="GRAVIDADE:N",
                y="UNIDADE:N",
                text=alt.Text("QTD:Q", format=".0f"),
                color=alt.value("#111"),
            )

            st.altair_chart((rects + labels).properties(height=340), use_container_width=True)
        else:
            st.info("Base sem colunas UNIDADE/GRAVIDADE.")

# ------------------ TABELAS EXTRAS ------------------
col_esq, col_dir = st.columns(2)

with col_esq:
    st.markdown('<div class="section">♻️ Reincidência por vistoriador (≥3)</div>', unsafe_allow_html=True)
    rec = (viewQ.groupby(["VISTORIADOR","ERRO"])["ERRO"]
           .size().reset_index(name="QTD").sort_values("QTD", ascending=False))
    rec = rec[rec["QTD"] >= 3]
    st.dataframe(rec, use_container_width=True, hide_index=True)

with col_dir:
    st.markdown('<div class="section">⚖️ Calibração por analista (% GG)</div>', unsafe_allow_html=True)
    if "ANALISTA" in viewQ.columns and "GRAVIDADE" in viewQ.columns:
        ana = (
            viewQ.assign(_gg=viewQ["GRAVIDADE"].isin(grav_gg).astype(int))
                 .groupby("ANALISTA")["_gg"]
                 .mean()
                 .reset_index(name="%GG")
        )
        # ordena do maior %GG para o menor
        ana = ana.sort_values("%GG", ascending=False)
        ana["%GG"] = (ana["%GG"] * 100).round(1)

        st.altair_chart(
            alt.Chart(ana).mark_bar().encode(
                x=alt.X("ANALISTA:N", axis=alt.Axis(labelAngle=0, labelLimit=180)),
                y=alt.Y("%GG:Q"),
                tooltip=["ANALISTA", alt.Tooltip("%GG:Q", format=".1f")]
            ).properties(height=340),
            use_container_width=True,
        )

st.markdown('<div class="section">📅 Erros por dia da semana</div>', unsafe_allow_html=True)
dow_map = {0:"Seg",1:"Ter",2:"Qua",3:"Qui",4:"Sex",5:"Sáb",6:"Dom"}
dow = pd.to_datetime(viewQ["DATA"], errors="coerce").dt.dayofweek.map(dow_map)
dow_counts = dow.value_counts().reindex(list(dow_map.values()), fill_value=0)
dow_df = pd.DataFrame({"DIA": dow_counts.index, "QTD": dow_counts.values})
if not dow_df.empty:
    st.altair_chart(bar_with_labels(dow_df, "DIA", "QTD", x_title="DIA DA SEMANA"),
                    use_container_width=True)


# ------------------ % ERRO (casamento com Produção) ------------------
st.markdown("---")
st.markdown('<div class="section">📐 % de erro por vistoriador</div>', unsafe_allow_html=True)

denom_mode = st.session_state.get("denom_mode_global", "Bruta (recomendado)")

# ============= METAS POR CIDADE – VELOX (GENÉRICO) =============
def _norm_city(x: str) -> str:
    return _strip_accents(_upper(x))

CITY_METAS = {
    _norm_city("ESTREITO"):     (5.0, 2.0),
    _norm_city("GRAJAÚ"):       (5.0, 2.0),
    _norm_city("IMPERATRIZ"):   (5.0, 2.0),
    _norm_city("PEDREIRAS"):    (5.0, 2.0),
    _norm_city("SÃO LUIS"):     (5.0, 2.0),
}

def _metas_cidade(cidade: str) -> tuple[float, float]:
    """Retorna (meta_erro_total, meta_erro_gg) para a cidade.
       Default: (3.5, 1.5) se não estiver no mapa."""
    return CITY_METAS.get(_norm_city(cidade), (3.5, 1.5))

TOL_AMARELO = 0.5  # tolerância em pontos percentuais

# ------------------ PRODUÇÃO COM FALLBACK ------------------
fallback_note = None

def _make_prod(df_prod):
    if df_prod.empty:
        return pd.DataFrame(columns=["VISTORIADOR", "vist", "rev", "liq"])
    out = (
        df_prod.groupby("VISTORIADOR", dropna=False)
               .agg(vist=("IS_REV", "size"),
                    rev=("IS_REV", "sum"))
               .reset_index()
    )
    out["liq"] = out["vist"] - out["rev"]
    return out

prod = _make_prod(viewP)

if prod["vist"].sum() == 0:
    if not dfP.empty:
        s_p_dates_all = pd.to_datetime(dfP["__DATA__"], errors="coerce").dt.date
        mask_mes_all = s_p_dates_all.map(
            lambda d: isinstance(d, date) and d.year == ref_year and d.month == ref_month
        )
        prod_month = dfP[mask_mes_all].copy()
        if "UNIDADE" in prod_month.columns and len(f_unids):
            prod_month = prod_month[prod_month["UNIDADE"].isin([_upper(u) for u in f_unids])]
        if "VISTORIADOR" in prod_month.columns and len(f_vists):
            prod_month = prod_month[prod_month["VISTORIADOR"].isin([_upper(v) for v in f_vists])]
        if set_vists_perfil is not None and "VISTORIADOR" in prod_month.columns:
            prod_month = prod_month[prod_month["VISTORIADOR"].isin(set_vists_perfil)]
        prod = _make_prod(prod_month)
        if prod["vist"].sum() > 0:
            fallback_note = "Usando produção do mês (fallback), pois não houve produção no período selecionado."

if prod["vist"].sum() == 0 and not dfP.empty:
    prod_all = dfP.copy()
    if set_vists_perfil is not None and "VISTORIADOR" in prod_all.columns:
        prod_all = prod_all[prod_all["VISTORIADOR"].isin(set_vists_perfil)]
    prod = _make_prod(prod_all)
    fallback_note = "Usando produção global (fallback), pois não há produção no mês/período selecionado."

# ------------------ QUALIDADE ------------------
qual = (
    viewQ.groupby("VISTORIADOR", dropna=False)
         .agg(erros=("ERRO", "size"),
              erros_gg=("GRAVIDADE", lambda s: s.isin(grav_gg).sum()))
         .reset_index()
)

# ------------------ CIDADE POR VISTORIADOR ------------------
# Usa UNIDADE da produção; se não tiver, cai para UNIDADE da qualidade
city_map = {}

if "UNIDADE" in viewP.columns and not viewP.empty:
    tmp = (
        viewP.groupby("VISTORIADOR")["UNIDADE"]
             .agg(lambda s: s.mode().iloc[0] if not s.mode().empty
                  else (s.dropna().iloc[0] if s.dropna().any() else ""))
    )
    city_map.update(tmp.to_dict())

if "UNIDADE" in viewQ.columns:
    tmp_q = (
        viewQ.groupby("VISTORIADOR")["UNIDADE"]
             .agg(lambda s: s.mode().iloc[0] if not s.mode().empty
                  else (s.dropna().iloc[0] if s.dropna().any() else ""))
    )
    for k, v in tmp_q.to_dict().items():
        city_map.setdefault(k, v)

# ------------------ BASE FINAL ------------------
base = prod.merge(qual, on="VISTORIADOR", how="outer").fillna(0)
base["CIDADE"] = base["VISTORIADOR"].map(city_map).fillna("")

den = base["liq"] if denom_mode.startswith("Líquida") else base["vist"]
den = den.replace({0: np.nan})

base["%ERRO"]    = ((base["erros"]    / den) * 100).round(1)
base["%ERRO_GG"] = ((base["erros_gg"] / den) * 100).round(1)

def _farol_pct(pct, cidade, meta_total, meta_gg, is_gg=False, tol=TOL_AMARELO):
    if pd.isna(pct):
        return "—"
    meta = meta_gg if is_gg else meta_total
    diff = pct - meta
    if diff <= 0:
        return "🟢"
    if diff <= tol:
        return "🟡"
    return "🔴"

# aplica farol usando meta (igual para todas as cidades)
farol_total = []
farol_gg    = []
for _, r in base.iterrows():
    mt_total, mt_gg = _metas_cidade(r.get("CIDADE", ""))
    farol_total.append(_farol_pct(r["%ERRO"],    r.get("CIDADE", ""), mt_total, mt_gg, is_gg=False))
    farol_gg.append(   _farol_pct(r["%ERRO_GG"], r.get("CIDADE", ""), mt_total, mt_gg, is_gg=True))

base["FAROL_%ERRO"]    = farol_total
base["FAROL_%ERRO_GG"] = farol_gg

# ------------------ FORMATAÇÃO E ORDENAÇÃO ------------------
fmt = base.copy()
for c in ["vist", "rev", "liq", "erros", "erros_gg"]:
    fmt[c] = pd.to_numeric(fmt[c], errors="coerce").fillna(0).astype(int)

def _fmt_val_pct(pct, emoji):
    if pd.isna(pct):
        return "—"
    return f"{emoji} {pct:.1f}%".replace(".", ",")

fmt["%ERRO"]    = fmt.apply(lambda r: _fmt_val_pct(r["%ERRO"],    r["FAROL_%ERRO"]), axis=1)
fmt["%ERRO_GG"] = fmt.apply(lambda r: _fmt_val_pct(r["%ERRO_GG"], r["FAROL_%ERRO_GG"]), axis=1)

# Ordenação decrescente pelo valor numérico real (%ERRO)
fmt_sorted = fmt.sort_values(
    by="%ERRO",
    key=lambda col: base.loc[col.index, "%ERRO"],
    ascending=False
)

cols_view = ["CIDADE", "VISTORIADOR", "vist", "rev", "liq", "erros", "erros_gg", "%ERRO", "%ERRO_GG"]

st.dataframe(
    fmt_sorted[cols_view],
    use_container_width=True,
    hide_index=True,
)

# ------------------ EXPORTAR EXCEL COM FAROL DE CORES ------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment
    ok_openpyxl = True
except Exception:
    ok_openpyxl = False

if not ok_openpyxl:
    st.warning("openpyxl não disponível — exportação colorida desativada.")
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Erros por Vistoriador"

    # Cabeçalho (com CIDADE)
    headers = ["CIDADE", "VISTORIADOR", "vist", "rev", "liq", "erros", "erros_gg", "%ERRO", "%ERRO_GG"]
    ws.append(headers)

    # Linhas
    for _, r in fmt_sorted.iterrows():
        ws.append([
            r.get("CIDADE", ""),
            r["VISTORIADOR"],
            int(r["vist"]), int(r["rev"]), int(r["liq"]),
            int(r["erros"]), int(r["erros_gg"]),
            r["%ERRO"], r["%ERRO_GG"],
        ])

    def _fill_from_farol(emoji: str) -> PatternFill:
        if isinstance(emoji, str) and "🟢" in emoji:
            return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        if isinstance(emoji, str) and "🟡" in emoji:
            return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        if isinstance(emoji, str) and "🔴" in emoji:
            return PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        return PatternFill(fill_type=None)

    # Aplicar cores nas colunas %ERRO (H) e %ERRO_GG (I)
    for i, (_, r) in enumerate(fmt_sorted.iterrows(), start=2):
        fill_total = _fill_from_farol(r.get("FAROL_%ERRO"))
        fill_gg    = _fill_from_farol(r.get("FAROL_%ERRO_GG"))

        ws[f"H{i}"].fill = fill_total
        ws[f"I{i}"].fill = fill_gg

        ws[f"H{i}"].alignment = Alignment(horizontal="center")
        ws[f"I{i}"].alignment = Alignment(horizontal="center")

    widths = {"A":18, "B":28, "C":10, "D":10, "E":10, "F":10, "G":10, "H":14, "I":14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    xbuf = io.BytesIO()
    wb.save(xbuf)
    xbuf.seek(0)

    st.download_button(
        label="📥 Baixar Excel com farol de cores",
        data=xbuf,
        file_name="erros_por_vistoriador_velox.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------ LEGENDA ------------------
with st.expander("Legenda do farol", expanded=False):
    st.write("🟢 Dentro da meta de qualidade")
    st.write(f"🟡 Até {TOL_AMARELO:.1f} pp acima da meta")
    st.write("🔴 Acima da meta + tolerância")
if fallback_note:
    st.caption(f"ℹ️ {fallback_note}")

# ------------------ TENDÊNCIA DE ERROS (projeção) ------------------
st.markdown("---")
st.markdown('<div class="section">📈 Tendência de erros (projeção até o fim do mês)</div>', unsafe_allow_html=True)

mtd = mtd_all.copy()
erros_mtd = (mtd.groupby("VISTORIADOR", dropna=False)["ERRO"]
             .size().reset_index(name="ERROS_MTD"))

ym_cur = f"{ref_year}-{ref_month:02d}"
metas_cur = dfMetas[dfMetas["YM"].fillna("").astype(str) == ym_cur].copy() if "YM" in dfMetas.columns else dfMetas.copy()
du_map = {}
if not metas_cur.empty and "DIAS_UTEIS" in metas_cur.columns:
    metas_cur["VISTORIADOR"] = metas_cur["VISTORIADOR"].astype(str).map(_upper)
    for _, r in metas_cur.iterrows():
        try:
            du_map[r["VISTORIADOR"]] = int(r["DIAS_UTEIS"]) if pd.notna(r["DIAS_UTEIS"]) else None
        except Exception:
            pass

rows = []
for _, r in erros_mtd.iterrows():
    v = r["VISTORIADOR"]; e_mtd = int(r["ERROS_MTD"])
    du_total = du_map.get(v, dias_totais_fallback) or dias_totais_fallback
    du_pass  = min(dias_passados, du_total) if du_total else dias_passados
    erros_dia = (e_mtd / du_pass) if du_pass else np.nan
    proj = int(round((erros_dia * du_total))) if not np.isnan(erros_dia) else e_mtd
    rows.append({
        "VISTORIADOR": v,
        "Erros (MTD)": e_mtd,
        "Erros/dia": round(erros_dia, 2) if not np.isnan(erros_dia) else 0.0,
        "Dias úteis passados": int(du_pass),
        "Dias úteis (mês)": int(du_total),
        "Projeção (mês)": proj
    })

if rows:
    tend_df = pd.DataFrame(rows).sort_values("Projeção (mês)", ascending=False)
    st.dataframe(tend_df, use_container_width=True, hide_index=True)
else:
    st.info("Sem dados de erros no mês/período para calcular a tendência.")

# ------------------ TABELA DETALHADA ------------------
if not fast_mode:
    st.markdown("---")
    st.markdown('<div class="section">🧾 Detalhamento (linhas da base)</div>', unsafe_allow_html=True)

    det = viewQ.copy()
    with st.expander("Filtros deste quadro (opcional)", expanded=False):
        c1, c2, c3 = st.columns(3)
        c4, c5, c6 = st.columns(3)

        _d = pd.to_datetime(det["DATA"], errors="coerce").dt.date
        _dmin, _dmax = (_d.min(), _d.max()) if _d.notna().any() else (date(2000,1,1), date(2000,1,1))
        f_data = c1.date_input("Data (início e fim)", value=(_dmin, _dmax), min_value=_dmin, max_value=_dmax, format="DD/MM/YYYY")

        f_placa = c2.text_input("Placa (contém)", "")

        opts_erro       = sorted(det["ERRO"].dropna().unique().tolist())
        opts_grav       = sorted(det["GRAVIDADE"].dropna().unique().tolist()) if "GRAVIDADE" in det.columns else []
        opts_cidade     = sorted(det["UNIDADE"].dropna().unique().tolist()) if "UNIDADE" in det.columns else []
        opts_vist       = sorted(det["VISTORIADOR"].dropna().unique().tolist()) if "VISTORIADOR" in det.columns else []
        opts_analista   = sorted(det["ANALISTA"].dropna().unique().tolist()) if "ANALISTA" in det.columns else []

        f_erros       = c3.multiselect("Erro", opts_erro, default=opts_erro)
        f_grav        = c4.multiselect("Gravidade", opts_grav, default=opts_grav)
        f_cidade      = c5.multiselect("Cidade / Unidade", opts_cidade, default=opts_cidade)
        f_vist        = c6.multiselect("Vistoriador", opts_vist, default=opts_vist)
        f_analista    = c6.multiselect("Analista", opts_analista, default=opts_analista, key="det_analista")

    if isinstance(f_data, tuple) and len(f_data) == 2:
        dini, dfim = f_data
        _d = pd.to_datetime(det["DATA"], errors="coerce").dt.date
        det = det[_d.between(dini, dfim)]

    if f_placa.strip():
        det = det[det["PLACA"].astype(str).str.contains(f_placa.strip(), case=False, na=False)]
    if len(f_erros):    det = det[det["ERRO"].isin(f_erros)]
    if len(f_grav):     det = det[det["GRAVIDADE"].isin(f_grav)]      if "GRAVIDADE"  in det.columns else det
    if len(f_cidade):   det = det[det["UNIDADE"].isin(f_cidade)]      if "UNIDADE"    in det.columns else det
    if len(f_vist):     det = det[det["VISTORIADOR"].isin(f_vist)]    if "VISTORIADOR" in det.columns else det
    if len(f_analista): det = det[det["ANALISTA"].isin(f_analista)]   if "ANALISTA"   in det.columns else det

    det_cols = ["DATA","UNIDADE","VISTORIADOR","PLACA","ERRO","GRAVIDADE","ANALISTA","OBS","TEMPO_CASA"]
    for c in det_cols:
        if c not in det.columns: det[c] = ""
    det = det[det_cols].sort_values(["DATA","UNIDADE","VISTORIADOR"])
    st.dataframe(det, use_container_width=True, hide_index=True)
    st.caption('<div class="table-note">* Filtros desta tabela são independentes dos filtros do topo do painel.</div>', unsafe_allow_html=True)

# ------------------ COMPARATIVO ATUAL x MÊS ANTERIOR (MESMO INTERVALO) ------------------
st.markdown("---")
st.markdown('<div class="section">📊 Comparativo por colaborador — período atual x mesmo período do mês anterior</div>', unsafe_allow_html=True)

prev_base = prev_base_cards  # já calculado com filtros
cur = (viewQ.groupby("VISTORIADOR", dropna=False)["ERRO"].size().reset_index(name="ERROS_ATUAL"))
prev = (prev_base.groupby("VISTORIADOR", dropna=False)["ERRO"].size().reset_index(name="ERROS_ANT"))

tab = cur.merge(prev, on="VISTORIADOR", how="outer").fillna(0)
tab["Δ"] = tab["ERROS_ATUAL"] - tab["ERROS_ANT"]
tab["VAR_%"] = np.where(tab["ERROS_ANT"] > 0, (tab["Δ"] / tab["ERROS_ANT"]) * 100, np.nan)

def _status(delta):
    if delta < 0: return "✅ Melhorou"
    if delta > 0: return "❌ Piorou"
    return "➡️ Igual"

tab["Status"] = tab["Δ"].map(_status)
tab_fmt = tab.copy()
tab_fmt["VAR_%"] = tab_fmt["VAR_%"].map(lambda x: "—" if pd.isna(x) else f"{x:.1f}%".replace(".", ","))

st.caption(
    f"Período atual: {periodo_atual_ini:%d/%m/%Y} – {periodo_atual_fim:%d/%m/%Y}  •  "
    f"Período anterior: {prev_ini:%d/%m/%Y} – {prev_fim:%d/%m/%Y}"
)
st.dataframe(
    tab_fmt.sort_values("ERROS_ATUAL", ascending=False)[
        ["VISTORIADOR","ERROS_ATUAL","ERROS_ANT","Δ","VAR_%","Status"]
    ],
    use_container_width=True, hide_index=True,
)

# ------------------ COMPARATIVO SEMANAL (2 a 4 semanas) ------------------
if not fast_mode:
    st.markdown("---")
    st.markdown("### 🔵 Comparativo semanal por vistoriador")

    def _clip_month(di, dfim):
        di = max(di, month_start)
        dfim = min(dfim, month_end)
        return di, dfim

    def _slice_q(df, di, dfim):
        d = pd.to_datetime(df["DATA"], errors="coerce").dt.date
        return df[d.between(di, dfim)]

    def _slice_p(df, di, dfim):
        d = pd.to_datetime(df["__DATA__"], errors="coerce").dt.date
        return df[d.between(di, dfim)]

    def _pct_week(qdf, pdf):
        """ERROS por vist. + %ERRO (bruta ou líquida) para uma janela semanal."""
        grav_gg = {"GRAVE", "GRAVISSIMO", "GRAVÍSSIMO"}

        if qdf.empty:
            qual = pd.DataFrame(columns=["VISTORIADOR","ERROS","ERROS_GG"])
        else:
            qual = (qdf.groupby("VISTORIADOR", dropna=False)
                    .agg(ERROS=("ERRO","size"),
                         ERROS_GG=("GRAVIDADE", lambda s: s.isin(grav_gg).sum()))
                    .reset_index())

        if pdf.empty:
            prod = pd.DataFrame(columns=["VISTORIADOR","vist","rev","liq"])
        else:
            prod = (pdf.groupby("VISTORIADOR", dropna=False)
                    .agg(vist=("IS_REV","size"), rev=("IS_REV","sum"))
                    .reset_index())
            prod["liq"] = prod["vist"] - prod["rev"]

        den_col = "liq" if denom_mode.startswith("Líquida") else "vist"
        out = prod.merge(qual, on="VISTORIADOR", how="outer").fillna(0)

        for c in ["vist","rev","liq","ERROS","ERROS_GG"]:
            if c in out.columns:
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)

        den = out[den_col].replace({0: np.nan}).astype(float)
        out["%ERRO"]    = (out["ERROS"]    / den * 100).round(1)
        out["%ERRO_GG"] = (out["ERROS_GG"] / den * 100).round(1)
        out["DEN"] = out[den_col].fillna(0).astype(int)

        return out[["VISTORIADOR","ERROS","%ERRO","ERROS_GG","%ERRO_GG","DEN"]]

    def _make_week_block(di, dfim, prefix, meta_list):
        q = _slice_q(viewQ, di, dfim)
        p = _slice_p(viewP, di, dfim)
        meta_list.append((prefix, di, dfim))
        return _pct_week(q, p).add_prefix(prefix)

    sem_fins = []
    cur_end = min(end_d, month_end)
    for _ in range(4):
        di = (pd.Timestamp(cur_end) - pd.Timedelta(days=6)).date()
        di, dfim = _clip_month(di, cur_end)
        if di > dfim or dfim < month_start:
            break
        sem_fins.append((di, dfim))
        cur_end = (pd.Timestamp(di) - pd.Timedelta(days=1)).date()
        if cur_end < month_start:
            break

    if len(sem_fins) < 2:
        st.info("Sem semanas suficientes no mês para montar o comparativo.")
    else:
        sem_fins = list(reversed(sem_fins))
        k = len(sem_fins)

        meta = []
        blocks = []
        for i, (di, dfim) in enumerate(sem_fins, start=1):
            blocks.append(_make_week_block(di, dfim, f"S{i}_", meta))

        from functools import reduce
        tab = reduce(
            lambda L, R: L.merge(R, left_on=f"{L.columns[0]}", right_on=f"{R.columns[0]}", how="outer"),
            blocks
        )

        def _pick_row(row):
            for i in range(1, k+1):
                v = row.get(f"S{i}_VISTORIADOR", "")
                if isinstance(v, str) and v.strip():
                    return v
            return ""
        tab["VISTORIADOR"] = tab.apply(_pick_row, axis=1)

        for c in tab.columns:
            if c.endswith("ERROS") or c.endswith("ERROS_GG") or c.endswith("DEN"):
                tab[c] = pd.to_numeric(tab[c], errors="coerce").fillna(0).astype(int)

        def _status_pp(delta):
            if pd.isna(delta): return "—"
            if delta < 0:     return f"Melhorou (↓ {abs(delta):.1f} pp)"
            if delta > 0:     return f"Piorou (↑ {delta:.1f} pp)"
            return "Sem alteração (↔)"

        for i in range(1, k):
            dcol = f"Δ_%ERRO_S{i}_S{i+1}"
            tab[dcol] = (tab[f"S{i+1}_%ERRO"] - tab[f"S{i}_%ERRO"]).round(1)
            tab[f"Status (S{i}→S{i+1})"] = tab[dcol].map(_status_pp)

        def _status3(p1, p2, p3):
            if any(pd.isna([p1, p2, p3])): return "—"
            d12 = p2 - p1; d23 = p3 - p2
            if d12 < 0 and d23 < 0: return "Continua melhorando (↓↓)"
            if d12 > 0 and d23 > 0: return "Continua piorando (↑↑)"
            if d12 < 0 and d23 > 0: return "Melhorou e depois piorou (↓↑)"
            if d12 > 0 and d23 < 0: return "Piorou e depois melhorou (↑↓)"
            return "Sem alteração (↔↔)"

        if k >= 3:
            tab["Status (3-semanas)"] = [
                _status3(r.get(f"S{k-2}_%ERRO", np.nan), r.get(f"S{k-1}_%ERRO", np.nan), r.get(f"S{k}_%ERRO", np.nan))
                for _, r in tab.iterrows()
            ]

        def _fmt_pct(x): return "—" if pd.isna(x) else f"{x:.1f}%".replace(".", ",")
        def _fmt_pp(x):  return "—" if pd.isna(x) else f"{x:.1f} pp".replace(".", ",")

        cols = ["VISTORIADOR"]
        for i in range(1, k+1):
            cols += [f"S{i}_ERROS", f"S{i}_%ERRO", f"S{i}_ERROS_GG", f"S{i}_%ERRO_GG"]

        for i in range(1, k):
            cols += [f"Δ_%ERRO_S{i}_S{i+1}", f"Status (S{i}→S{i+1})"]

        if k >= 3:
            cols += ["Status (3-semanas)"]

        out = tab[cols].copy()

        for c in out.columns:
            if c.endswith("%ERRO") or c.endswith("%ERRO_GG"):
                out[c] = out[c].map(_fmt_pct)
            elif c.startswith("Δ_%ERRO_"):
                out[c] = out[c].map(_fmt_pp)

        order_key = tab[f"S{k}_%ERRO"].fillna(-1).values
        out = out.iloc[np.argsort(-order_key)]

        legend_parts = []
        for i, (prefix, di, dfim) in enumerate(meta, start=1):
            label = f"Semana {i}: {di:%d/%m}–{dfim:%d/%m}"
            if i == k:
                label = label.replace(f"Semana {i}", f"Semana {i} (atual)")
            legend_parts.append(label)
        st.caption("  ·  ".join(legend_parts))

        st.dataframe(out.reset_index(drop=True), use_container_width=True, hide_index=True)

# ------------------ RANKINGS ------------------
st.markdown("---")
st.markdown('<div class="section">🏁 Top 5 melhores × piores (por % de erro)</div>', unsafe_allow_html=True)

rank = (base.copy())
rank = rank[den > 0].replace({np.inf: np.nan}).dropna(subset=["%ERRO"])

den_col = "liq" if denom_mode.startswith("Líquida") else "vist"
col_titulo_den = "vistorias líquidas" if den_col == "liq" else "vistorias"
cols_rank = ["VISTORIADOR", den_col, "erros", "%ERRO", "%ERRO_GG"]
rank_view = rank[cols_rank].rename(columns={den_col: col_titulo_den})

for c in [col_titulo_den, "erros"]:
    if c in rank_view.columns: rank_view[c] = rank_view[c].astype(int)
for c in ["%ERRO", "%ERRO_GG"]:
    if c in rank_view.columns: rank_view[c] = rank_view[c].map(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")

c_best, c_worst = st.columns(2)
with c_best:
    best5  = rank_view.sort_values("%ERRO", ascending=True).head(5)
    st.subheader("🏆 Top 5 melhores (menor %Erro)")
    st.dataframe(best5.reset_index(drop=True), use_container_width=True, hide_index=True)
with c_worst:
    worst5 = rank_view.sort_values("%ERRO", ascending=False).head(5)
    st.subheader("⚠️ Top 5 piores (maior %Erro)")
    st.dataframe(worst5.reset_index(drop=True), use_container_width=True, hide_index=True)

# ------------------ FRAUDE ------------------
st.markdown("---")
st.markdown('<div class="section">🚨 Tentativa de Fraude — Detalhamento</div>', unsafe_allow_html=True)
fraude_mask = viewQ["ERRO"].astype(str).str.upper().str.contains(r"\bTENTATIVA DE FRAUDE\b", na=False)
df_fraude = viewQ[fraude_mask].copy()
if df_fraude.empty:
    st.info("Nenhum registro de Tentativa de Fraude no período/filtros selecionados.")
else:
    cols_fraude = ["DATA","UNIDADE","VISTORIADOR","PLACA","ERRO","GRAVIDADE","ANALISTA","OBS","TEMPO_CASA"]
    for c in cols_fraude:
        if c not in df_fraude.columns: df_fraude[c] = ""
    df_fraude = df_fraude[cols_fraude].sort_values(["DATA","UNIDADE","VISTORIADOR"])
    st.dataframe(df_fraude, use_container_width=True, hide_index=True)
    st.caption('<div class="table-note">* Somente linhas cujo ERRO é exatamente “TENTATIVA DE FRAUDE”.</div>', unsafe_allow_html=True)

# ------------------ HISTÓRICO BOTTOM 5 (últimos 3 meses) ------------------
st.markdown("---")
st.markdown(
    '<div class="section">📚 Histórico dos Bottom 5 (últimos 3 meses)</div>',
    unsafe_allow_html=True
)

# Nomes dos 5 piores já calculados acima (worst5)
bottom_names = []
try:
    bottom_names = worst5["VISTORIADOR"].astype(str).tolist()
except Exception:
    bottom_names = []

if not bottom_names:
    st.info("Ainda não há vistoriadores no ranking de piores para montar o histórico.")
else:
    # ym_all = lista de meses disponíveis no formato 'AAAA-MM'
    # ym_sel  = mês atual selecionado no topo
    try:
        idx_cur = ym_all.index(ym_sel)
    except ValueError:
        idx_cur = len(ym_all) - 1

    # Últimos 3 meses: atual + até 2 anteriores
    ini = max(0, idx_cur - 2)
    meses_janela = ym_all[ini: idx_cur + 1]

    if not meses_janela:
        st.info("Não há meses suficientes na base para montar o histórico.")
    else:
        # Base com nomes dos bottom 5 (do mês atual)
        hist_df = pd.DataFrame({"VISTORIADOR": sorted(set(bottom_names))})

        labels_legenda = []

        for ym in meses_janela:
            ano = int(ym[:4])
            mes = int(ym[5:7])
            label_mes = f"{mes:02d}/{ano}"
            labels_legenda.append(label_mes)

            # --------- BASE DE QUALIDADE E PRODUÇÃO POR MÊS ---------
            if ym == ym_sel:
                # MÊS ATUAL → usa exatamente o mesmo recorte do painel
                dq_m = viewQ.copy()
                dp_m = viewP.copy()
            else:
                # MÊS ANTERIOR → mês cheio, com mesmos filtros de unidade/perfil
                # Qualidade
                dq_m = dfQ.copy()
                dt_q = pd.to_datetime(dq_m["DATA"], errors="coerce")
                mask_mq = (dt_q.dt.year.eq(ano) & dt_q.dt.month.eq(mes))
                dq_m = dq_m[mask_mq].copy()

                if len(f_unids) and "UNIDADE" in dq_m.columns:
                    dq_m = dq_m[dq_m["UNIDADE"].isin([_upper(u) for u in f_unids])]
                if "TEMPO_CASA" in dq_m.columns and perfil_sel != "Todos":
                    alvo = "NOVATO" if perfil_sel == "Novatos" else "VETERANO"
                    dq_m = dq_m[dq_m["TEMPO_CASA"] == alvo]

                # Produção
                if not dfP.empty:
                    dp_m = dfP.copy()
                    dt_p = pd.to_datetime(dp_m["__DATA__"], errors="coerce")
                    mask_mp = (dt_p.dt.year.eq(ano) & dt_p.dt.month.eq(mes))
                    dp_m = dp_m[mask_mp].copy()

                    if len(f_unids) and "UNIDADE" in dp_m.columns:
                        dp_m = dp_m[dp_m["UNIDADE"].isin([_upper(u) for u in f_unids])]
                    if set_vists_perfil is not None and "VISTORIADOR" in dp_m.columns:
                        dp_m = dp_m[dp_m["VISTORIADOR"].isin(set_vists_perfil)]
                else:
                    dp_m = dfP.copy()

            # Produção agrupada (reaproveita função _make_prod)
            prod_m = _make_prod(dp_m)

            # Qualidade agrupada
            if dq_m.empty:
                qual_m = pd.DataFrame(columns=["VISTORIADOR", "erros", "erros_gg"])
            else:
                qual_m = (
                    dq_m.groupby("VISTORIADOR", dropna=False)
                        .agg(
                            erros=("ERRO", "size"),
                            erros_gg=("GRAVIDADE", lambda s: s.isin(grav_gg).sum())
                        )
                        .reset_index()
                )

            # Junta produção + qualidade
            base_m = prod_m.merge(qual_m, on="VISTORIADOR", how="outer").fillna(0)

            # Calcula %ERRO e %ERRO_GG usando o mesmo denominador (bruta ou líquida)
            den_hist = base_m["liq"] if denom_mode.startswith("Líquida") else base_m["vist"]
            den_hist = den_hist.replace({0: np.nan})

            base_m["%ERRO"] = ((base_m["erros"]    / den_hist) * 100).round(1)
            base_m["%ERRO_GG"] = ((base_m["erros_gg"] / den_hist) * 100).round(1)

            # Ranking do mês para saber quem foi bottom 5 naquele mês (por %ERRO total)
            rank_m = base_m.copy()
            rank_m = rank_m[den_hist > 0].replace({np.inf: np.nan}).dropna(subset=["%ERRO"])
            rank_m = rank_m.sort_values("%ERRO", ascending=False)
            bottom_m = rank_m["VISTORIADOR"].astype(str).head(5).tolist()

            # Foca só nos bottom atuais (bottom_names)
            tmp = base_m[["VISTORIADOR", "%ERRO", "%ERRO_GG"]].copy()
            tmp["VISTORIADOR"] = tmp["VISTORIADOR"].astype(str)
            tmp = tmp[tmp["VISTORIADOR"].isin(bottom_names)]

            tmp = tmp.rename(columns={
                "%ERRO":    f"%Erro {label_mes}",
                "%ERRO_GG": f"%Erro GG {label_mes}",
            })
            tmp[f"Bottom {label_mes}"] = tmp["VISTORIADOR"].isin(bottom_m)

            hist_df = hist_df.merge(tmp, on="VISTORIADOR", how="left")

        # --- GARANTIR QUE O MÊS ATUAL USE EXATAMENTE O TOP 5 DO PAINEL ---
        if labels_legenda:
            col_bottom_cur = f"Bottom {labels_legenda[-1]}"
            if col_bottom_cur in hist_df.columns:
                nomes_bottom_atual = [str(v) for v in bottom_names]
                hist_df[col_bottom_cur] = hist_df["VISTORIADOR"].astype(str).isin(nomes_bottom_atual)

        # Calcula quantos meses cada um apareceu no bottom
        bottom_cols = [c for c in hist_df.columns if c.startswith("Bottom ")]
        if bottom_cols:
            hist_df[bottom_cols] = hist_df[bottom_cols].fillna(False)
            hist_df["Meses no bottom"] = hist_df[bottom_cols].sum(axis=1)

            def _icone_reinc(x):
                if x >= 3:
                    return "🔥 3 meses no bottom"
                if x == 2:
                    return "⚠️ 2 meses no bottom"
                if x == 1:
                    return "🆕 Entrou agora"
                return "✅ Saiu do bottom"

            hist_df["Situação"] = hist_df["Meses no bottom"].map(_icone_reinc)
        else:
            hist_df["Meses no bottom"] = 0
            hist_df["Situação"] = "—"

        # Formata %Erro e flags de bottom para ficar mais visual
        for c in hist_df.columns:
            if c.startswith("%Erro GG "):
                hist_df[c] = hist_df[c].map(
                    lambda x: "—" if pd.isna(x) else f"{float(x):.1f}%".replace(".", ",")
                )
            elif c.startswith("%Erro "):
                hist_df[c] = hist_df[c].map(
                    lambda x: "—" if pd.isna(x) else f"{float(x):.1f}%".replace(".", ",")
                )
            if c.startswith("Bottom "):
                hist_df[c] = hist_df[c].map(lambda v: "🔴" if bool(v) else "—")

        # Ordena pelos mais reincidentes / pior %Erro atual
        col_pct_atual = f"%Erro {labels_legenda[-1]}"
        if col_pct_atual in hist_df.columns:
            order_key = hist_df["Meses no bottom"] * 1000
            try:
                num_pct = (
                    hist_df[col_pct_atual]
                    .astype(str)
                    .str.replace("%", "")
                    .str.replace(",", ".")
                    .astype(float)
                )
                order_key = order_key + num_pct
            except Exception:
                pass
            hist_df = hist_df.iloc[np.argsort(-order_key)].reset_index(drop=True)

        # Colunas na ordem: nome, situação, reincidência, depois meses
        cols_show = ["VISTORIADOR", "Situação", "Meses no bottom"]
        for label_mes in labels_legenda:
            pct_col   = f"%Erro {label_mes}"
            pctgg_col = f"%Erro GG {label_mes}"
            btm_col   = f"Bottom {label_mes}"
            if pct_col in hist_df.columns:
                cols_show.append(pct_col)
            if pctgg_col in hist_df.columns:
                cols_show.append(pctgg_col)
            if btm_col in hist_df.columns:
                cols_show.append(btm_col)

        out_hist = hist_df[cols_show].copy()

        st.dataframe(
            out_hist,
            use_container_width=True,
            hide_index=True,
        )

        legenda_txt = " · ".join(
            [f"{lab}: %Erro, %Erro GG e 🔴 se ficou entre os 5 piores no mês" for lab in labels_legenda]
        )
        st.caption(
            "Coluna **Situação** mostra a reincidência dos 5 piores do mês atual nos últimos meses. "
            + legenda_txt
        )

